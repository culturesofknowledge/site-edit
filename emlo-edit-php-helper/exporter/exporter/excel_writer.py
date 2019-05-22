import csv

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.compat import basestring

import codecs
# from io import open


def unicode_csv_reader(unicode_csv_data, dialect=csv.excel, **kwargs):
	# csv.py doesn't do Unicode; encode temporarily as UTF-8:
	csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
		dialect=dialect, **kwargs)
	for row in csv_reader:
		# decode UTF-8 back to Unicode, cell by cell:
		yield [unicode(cell, 'utf-8') for cell in row]


def utf_8_encoder(unicode_csv_data):
	for line in unicode_csv_data:
		yield line.encode('utf-8')


class ExcelWriter:

	def convert(self, settings, skip_styles=False ):
		"""
		Add a list of CSVs to a new Excel file
		:param settings= {
			'sheets' : [
				{
					'filelocation': '<string csv filename>',
					'sheetname': '<string sheet name>',
					'has_titles': '<bool default:true>',  # Use bold if there are titles.
				}
			],
			'outputname' : '<output xlsx name>'
		}
		:return: None
		"""

		wb = Workbook()

		if not skip_styles :
			col_title_font = Font( size=11, italic=True )
			col_font = Font( size=11 )

		first = True
		for sheet in settings["sheets"]:

			if first:
				ws = wb.active
				first = False
			else :
				ws = wb.create_sheet()

			ws.title = sheet["sheetname"]

			with codecs.open(sheet["filelocation"], "r", encoding="utf-8") as f:
				csv_reader = unicode_csv_reader(f)
				row_count = 1

				for csv_row in csv_reader:
					col_count = 1

					#if row_count % 10 == 0 :
					#	print( sheet["sheetname"] + " : " + str(row_count) )

					for cell in csv_row :


						c = ws.cell(row=row_count, column=col_count)
						if not skip_styles :
							if row_count == 1 and sheet.get( "has_titles", True ) :
								# Set bold if titles
								c.font = col_title_font
							else :
								c.font = col_font

						c.value = cell.replace( u'\u000B', '' )

						col_count += 1

					row_count += 1

			if not skip_styles :
				# Adjust widths of columns
				dims = {}
				for row in ws.rows:
					for cell in row:
						if cell.value:
							dims[cell.column] = max((dims.get(cell.column, 0), self.calculate_width(cell.value)))

				#for col, value in dims.items():
				#	print( value )
				#	ws.column_dimensions[col].width = "" + str(value)  # TODO: work out why this is broken...

			wb.save(settings["outputname"])

	@staticmethod
	def calculate_width( value ):
		""" Estimate the width of a cell based on its string """
		width = 0
		max_line_width = 0
		for c in value :
			if c in "\n" :
				if width > max_line_width :
					width = max_line_width
				width = 0
			else:	
				if c in "ijlIJ(){}[]. " :
					width += 0.6
				elif c in "mwMW" :
					width += 1.6
				elif c in "ABCDEFGHKLNOPQRSTUVXYZ" :
					width += 1.2
				else :
					width += 1.1

		if max_line_width > width :
			width = max_line_width

		return max(5 , min(50, int( round(width) )) )

if __name__ == "__main__":

	ew = ExcelWriter()

	data = {
		"sheets" : [
			{
				"filelocation" : "Test/work.csv",
				"sheetname" : "Works"
			},
			{
				"filelocation" : "Test/person.csv",
				"sheetname" : "People",
				"has_titles" : False
			},
			{
				"filelocation" : "Test/location.csv",
				"sheetname" : "Locations",
				"has_titles" : False
			}
		],
		"outputname" : "Test/test1.xlsx"
	}

	# filenames = [
	# 	'archives',
	# 	'correspondents',
	# 	'letter',
	# 	'msrefs',
	# 	'mswitnesses',
	# 	'namestrings',
	# 	'peopleidx',
	# 	'peoplementioned',
	# 	'placenames',
	# 	'places',
	# 	'placesmentioned',
	# 	'printrefs',
	# 	'printwitnesses',
	# 	'subjectidx',
	# 	'subjects'
	# ]
	#
	# location = '/home/matthew/Projects/emlo/bacon/bacon_'
	#
	# data = {
	# 	"sheets" : [],
	# 	"outputname" : location + "all.xlsx"
	# }
	#
	# for name in filenames :
	# 	data["sheets"].append( {
	# 		"filelocation" : location + name + ".csv",
	# 		"sheetname" : name
	# 	} )


	ew.convert( data )

