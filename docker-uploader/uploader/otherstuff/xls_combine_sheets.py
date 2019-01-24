from __future__ import print_function

import sys, argparse, os.path, fnmatch
import xlrd
from openpyxl import Workbook

def create_csvs( filename ) :

	workbook = xlrd.open_workbook( filename )

	wb = Workbook()
	worksheet_all = wb.create_sheet()
	worksheet_all.title = "All"

	sheet_count = 1
	row_count_all = 1  # different libraries use different start values...
	while sheet_count <= 81 :

		sheet_name = "Sheet" + str(sheet_count)

		worksheet = None
		try:
			worksheet = workbook.sheet_by_name( sheet_name )
		except xlrd.biffh.XLRDError:
			pass

		if worksheet is None :
			print( "Can't find worksheet '" + sheet_name )

		#
		# Open CSV writer
		#

		if worksheet is not None :

			for row in xrange( 0, worksheet.nrows ):

				for col in xrange( 0, worksheet.ncols ):

					print( row_count_all, row, col )

					value = worksheet.cell_value( row, col )
					#type = worksheet.cell_type(row, col )
					#print( type, value )

					cell_to = worksheet_all.cell( row_count_all, col+1 )
					cell_to.value = value

				row_count_all += 1

			print ("Extracted " + str(worksheet.nrows) + " rows from worksheet " + sheet_name )

		sheet_count += 1

	wb.save( "combined.xlsx")


if __name__ == "__main__":

	create_csvs( "md-aihs.xlsx" )