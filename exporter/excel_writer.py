import csv

from openpyxl import Workbook
from openpyxl.styles import Font

# from io import open

class ExcelWriter:

	def convert(self, settings ):
		"""
		Add a list of CSVs to a new Excel file
		:param settings: {
			sheets : [
				{
					filelocation: <string csv filename>,
					sheetname: <string sheet name>,
					has_titles: <bool default:true>
				}
			]
			outputname : <output xlsx name
		}
		:return: None
		"""

		wb = Workbook()
		col_title_font = Font( size=11, italic=True )
		col_font = Font( size=11 )

		first = True
		for sheet in settings["sheets"]:

			if first:
				ws = wb.active
				first = False
			else :
				ws = wb.create_sheet()

			ws.title = sheet["sheetname"]

			with open(sheet["filelocation"], "r" ) as f:  #encoding="utf-8") as f:
				csv_reader = csv.reader(f)
				row_count = 1

				for csv_row in csv_reader:
					col_count = 1
					for cell in csv_row :

						c = ws.cell(row=row_count, column=col_count)

						if row_count == 1 and sheet.get( "has_titles", True ) :
							# Set bold if titles
							c.font = col_title_font
						else :
							c.font = col_font

						c.value = cell

						col_count += 1

					row_count += 1

			# Adjust widths of columns
			dims = {}
			for row in ws.rows:
				for cell in row:
					if cell.value:
						dims[cell.column] = max((dims.get(cell.column, 0), self.calculate_width(cell.value)))
			
			for col, value in dims.items():
				ws.column_dimensions[col].width = value

			wb.save(settings["outputname"])

	@staticmethod
	def calculate_width( value ):
		""" Estimate the width of a cell based on its string """
		width = 0
		max_line_width = 0
		for c in value :
			if c in "\n" :
				if width > max_line_width :
					width = max_line_width
				width = 0
			else:	
				if c in "ijlIJ(){}[]. " :
					width += 0.7
				elif c in "mwMWQ" :
					width += 1.8
				else :
					width += 1.2

		if max_line_width > width :
			width = max_line_width

		return max(5 , width)

if __name__ == "__main__":

	ew = ExcelWriter()
	ew.convert( {
		"sheets" : [
			{
				"filelocation" : "Test/work.csv",
				"sheetname" : "Works"
			},
			{
				"filelocation" : "Test/person.csv",
				"sheetname" : "People",
				"has_titles" : False
			},
			{
				"filelocation" : "Test/location.csv",
				"sheetname" : "Locations",
				"has_titles" : False
			}
		],
		"outputname" : "Test/test1.xlsx"
	})

